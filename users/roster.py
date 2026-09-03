"""Ведомость: выгрузка ответов гугл-формы, из которой курс заводят разом.

Ни одного столбца в этом файле мы не заказывали — шапку гугл ставит из текста вопросов,
а вопросы пишет человек. Поэтому столбец ищется по НАЧАЛУ СЛОВА в шапке: «Удобная почта»,
«Почта» и «e-mail» находятся одинаково, как и «Группа» с «Номером группы», а лишние
столбцы («Интересный факт», «Отметка времени») не находятся ничем и проходят мимо.

Форма живёт здесь же, а не в forms.py: она вся — одно поле вокруг разбора ниже,
и порознь их читать незачем.
"""

import re
from datetime import date, datetime

from django import forms
from django.core.exceptions import ValidationError
from django.core.validators import validate_email

from openpyxl import load_workbook

from core.models import Team

from .forms import handle
from .models import User

MAX_FILE = 5 * 1024 * 1024
MAX_ROWS = 500  # это ведомость курса, а не выгрузка всего университета

# Что за столбец — по началу слова в шапке. Началом, а не целым словом: спрашивают
# и «Группа», и «Номер группы», и «Ваша группа» — падеж у каждой формы свой.
COLUMNS = {
    "fio": ("фио",),
    "email": ("почт", "email", "mail"),
    "team": ("групп",),
    "birthday": ("рожд",),
    "tg": ("телеграм", "telegram"),
}
# Эти — только целиком: «др» началом сидит в «друге», а «тг» — в «тгк».
EXACT = {"birthday": ("др",), "tg": ("тг",)}
# Без этих трёх ведомость бесполезна, остальное — как получится.
REQUIRED = {"fio": "ФИО", "email": "почта", "team": "группа"}

# Кириллица и латиница в номере группы неразличимы на глаз, и в жизни встречаются обе:
# в базе магистратура записана латинской M, а в ведомости от 03.09.2026 в одном и том же
# столбце соседствовали «М07-601» кириллицей и «M07-601» латиницей. Отказ по такой причине
# человеку не объяснить — поэтому обе стороны приводим к латинице.
LOOKALIKE = str.maketrans("АВЕКМНОРСТУХ", "ABEKMHOPCTYX")

# Форматы даты на случай, когда столбец в форме был текстовым: у даты из настоящей
# ячейки тип уже правильный, её формат openpyxl читает сам.
DATES = ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d")


def _words(cell):
    """Шапка словами: «Удобная почта» → {удобная, почта}.

    Именно словами, а не поиском подстроки: «др» подстрокой сидит и в «адресе»,
    и столбец с адресом уехал бы в день рождения.
    """
    return set(re.findall(r"[^\W\d_]+", str(cell or "").lower()))


def _columns(header):
    """{поле: номер столбца}. Первое совпадение выигрывает: если про почту в форме
    спрашивали дважды, берём левый столбец — он и есть основной."""
    found = {}
    for index, cell in enumerate(header):
        words = _words(cell)
        for field, stems in COLUMNS.items():
            if field in found:
                continue
            if words.intersection(EXACT.get(field, ())) or any(word.startswith(stems) for word in words):
                found[field] = index
    return found


def _cell(row, at, field):
    index = at.get(field)
    return row[index] if index is not None and index < len(row) else None


def _canon(number):
    return str(number or "").strip().upper().translate(LOOKALIKE)


def _fio(value):
    """«Фамилия Имя Отчество» → три поля, или None, если слов меньше двух.

    Всё после второго слова считаем отчеством: двойное отчество целиком лучше,
    чем обрезанное. Отчество бывает не у всех, и его отсутствие — не ошибка.
    """
    parts = str(value or "").split()
    if len(parts) < 2:
        return None
    return parts[0], parts[1], " ".join(parts[2:])


def _birthday(value):
    """День рождения. Не разобрали — оставляем пусто и не спорим: поле необязательное,
    и валить из-за него регистрацию всего курса не за что."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for shape in DATES:
        try:
            return datetime.strptime(str(value or "").strip(), shape).date()
        except ValueError:
            continue
    return None


def read(upload):
    """Разобрать ведомость: (люди, проблемы).

    Люди — несохранённые `User`, проблемы — строки для того, кто грузил файл. Пока
    список проблем не пуст, не заводят никого: половина курса, заведённая молча, хуже,
    чем ни одного, — узнать потом, кто не попал, будет неоткуда.
    """
    upload.seek(0)
    try:
        book = load_workbook(upload, read_only=True, data_only=True)
        sheet = book.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
    except Exception:  # noqa: BLE001 — файл принесли снаружи, сломаться он может как угодно
        return [], ["Не получилось прочитать файл — нужна таблица .xlsx"]

    if header is None:
        return [], ["В файле нет ни одной строки"]
    at = _columns(header)
    if missing := [title for field, title in REQUIRED.items() if field not in at]:
        return [], [f"В шапке не нашлось столбцов: {', '.join(missing)}"]

    # Все группы разом: их два десятка, а сравнение всё равно идёт по канону —
    # запросом по номеру мы бы промахнулись мимо чужой буквы (см. LOOKALIKE).
    teams = {_canon(team.number): team for team in Team.objects.all()}
    people, problems, seen = [], [], set()

    for line, row in enumerate(rows, start=2):
        if not any(str(cell or "").strip() for cell in row):
            continue
        if line - 1 > MAX_ROWS:
            problems.append(f"В файле больше {MAX_ROWS} строк — это не похоже на ведомость курса")
            break

        fio = _fio(_cell(row, at, "fio"))
        if fio is None:
            problems.append(f"строка {line}: в ФИО меньше двух слов")
            continue
        if max(len(part) for part in fio) > 50:
            problems.append(f"строка {line}: в ФИО слово длиннее 50 букв")
            continue

        email = str(_cell(row, at, "email") or "").strip()
        try:
            validate_email(email)
        except ValidationError:
            problems.append(f"строка {line}: «{email}» не похоже на почту")
            continue
        if email.lower() in seen:
            problems.append(f"строка {line}: почта {email} в файле уже встречалась")
            continue
        seen.add(email.lower())

        group = str(_cell(row, at, "team") or "").strip()
        team = teams.get(_canon(group))
        if team is None:
            problems.append(f"строка {line}: группы «{group}» нет в базе")
            continue

        surname, name, patronymic = fio
        people.append(User(
            surname=surname, name=name, patronymic=patronymic, email=email, team=team,
            birthday=_birthday(_cell(row, at, "birthday")),
            # Телеграм режем молча: поле необязательное и украшательское, и отказывать
            # из-за него всему курсу не за что.
            tg_page=handle(str(_cell(row, at, "tg") or ""), "t.me")[:50],
        ))

    return people, problems


class RosterForm(forms.Form):
    file = forms.FileField(label="Файл с ответами")

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.people = []

    def clean_file(self):
        upload = self.cleaned_data["file"]
        if upload.size > MAX_FILE:
            raise ValidationError(f"Файл тяжелее {MAX_FILE // 1024 // 1024} МБ — это точно не ведомость")
        return upload

    def clean(self):
        data = super().clean()
        if upload := data.get("file"):
            self.people, problems = read(upload)
            if problems:
                # Списком, а не одной строкой: проблем бывает столько же, сколько строк,
                # и поправить их человеку нужно все разом, а не по одной за загрузку.
                raise ValidationError(problems)
        return data
